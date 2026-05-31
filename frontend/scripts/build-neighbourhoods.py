"""Build the enriched neighbourhood choropleth dataset.

Merges Toronto's 158-neighbourhood polygons with 2021 Census attributes and the
2021 Ontario Marginalization Index, computes geodesic area + density + a set of
percentage metrics, and emits a compact GeoJSON the web app paints as choropleths.

Raw data lives outside the repo (the `data/` folder is git-ignored), so this
reads from an absolute DATA_DIR and writes the *derived* artifact into
`frontend/public/neighbourhoods.json`, which IS committed. Run from anywhere:

    python frontend/scripts/build-neighbourhoods.py
    DATA_DIR=/abs/path/to/data python frontend/scripts/build-neighbourhoods.py

Output feature properties (per neighbourhood):
    num, name, area_km2, pop, density,
    low_income_pct, transit_commute_pct, car_pct, active_pct,
    senior_pct, renter_pct,
    noc0_pct .. noc9_pct                 (occupation share of labour force)
    marg_material, marg_age_labour, marg_households, marg_racialized  (1-5 quintiles)
"""
import json, math, os, re
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(HERE, "..", ".."))
DATA = os.environ.get("DATA_DIR", os.path.join(REPO, "data"))
GEO = os.path.join(DATA, "geospatial", "neighbourhoods-158.geojson")
PROF = os.path.join(DATA, "census-demographics", "neighbourhood-profiles-2021.xlsx")
MARG = os.path.join(DATA, "census-demographics", "on-marg-2021-toronto-n158.xlsx")
OUT = os.path.join(HERE, "..", "public", "neighbourhoods.json")

R = 6378137.0  # WGS84 equatorial radius (m)


# --- geodesic polygon area (lon/lat rings) --------------------------------
def ring_area(ring):
    n = len(ring)
    if n < 3:
        return 0.0
    total = 0.0
    for i in range(n):
        lon1, lat1 = ring[i][0], ring[i][1]
        lon2, lat2 = ring[(i + 1) % n][0], ring[(i + 1) % n][1]
        total += math.radians(lon2 - lon1) * (
            2 + math.sin(math.radians(lat1)) + math.sin(math.radians(lat2))
        )
    return total * R * R / 2.0


def polygon_area_km2(coords):
    a = abs(ring_area(coords[0]))
    for hole in coords[1:]:
        a -= abs(ring_area(hole))
    return a / 1e6


def geom_area_km2(geom):
    if geom["type"] == "Polygon":
        return polygon_area_km2(geom["coordinates"])
    if geom["type"] == "MultiPolygon":
        return sum(polygon_area_km2(p) for p in geom["coordinates"])
    return 0.0


def round_geom(geom, nd=5):
    def rc(ring):
        return [[round(x, nd), round(y, nd)] for x, y in ring]
    if geom["type"] == "Polygon":
        geom["coordinates"] = [rc(r) for r in geom["coordinates"]]
    elif geom["type"] == "MultiPolygon":
        geom["coordinates"] = [[rc(r) for r in poly] for poly in geom["coordinates"]]
    return geom


# --- census profile parse --------------------------------------------------
# Layout: col 0 = variable label, row 2 = neighbourhood numbers, cols 1..158 =
# neighbourhood values. We locate rows by label (resilient to row shifts) and
# scope sub-category lookups to a window after their section header.
wb = load_workbook(PROF, data_only=True, read_only=True)
ws = wb["hd2021_census_profile"]
rows = list(ws.iter_rows(values_only=True))
numbers = rows[1]
col_num = {
    ci: int(numbers[ci])
    for ci in range(1, len(numbers))
    if isinstance(numbers[ci], (int, float))
}
labels = [str(r[0]).strip() if r and r[0] is not None else "" for r in rows]


def row_find(pred, start=0, end=None):
    end = end if end is not None else len(rows)
    for i in range(start, min(end, len(rows))):
        if pred(labels[i]):
            return i
    return -1


def vals(i):
    """Map neighbourhood number -> numeric value for row i (empty if i<0)."""
    if i < 0:
        return {}
    r = rows[i]
    return {n: r[ci] for ci, n in col_num.items() if isinstance(r[ci], (int, float))}


i_pop = row_find(lambda l: l.startswith("Total - Age groups of the population"))
pop = vals(i_pop)
senior = vals(row_find(lambda l: l == "65 years and over", i_pop, i_pop + 60))

i_ten = row_find(lambda l: l.startswith("Total - Private households by tenure"))
ten_total = vals(i_ten)
renter = vals(row_find(lambda l: l == "Renter", i_ten, i_ten + 6))

i_li = row_find(lambda l: "low income" in l.lower() and "lim-at" in l.lower() and "%" in l)
low_income = vals(i_li)  # already a percentage

i_cm = row_find(lambda l: l.startswith("Total - Main mode of commuting"))
cm_total = vals(i_cm)
transit = vals(row_find(lambda l: l == "Public transit", i_cm, i_cm + 14))
car = vals(row_find(lambda l: l.startswith("Car, truck or van") and "as a" not in l, i_cm, i_cm + 14))
walked = vals(row_find(lambda l: l == "Walked", i_cm, i_cm + 14))
bicycle = vals(row_find(lambda l: l == "Bicycle", i_cm, i_cm + 14))

i_occ = row_find(lambda l: l == "All occupations")
all_occ = vals(i_occ)
noc = {d: vals(row_find(lambda l, d=d: re.match(rf"^{d}\s", l), i_occ, i_occ + 14)) for d in range(10)}


# --- marginalization parse -------------------------------------------------
marg = {}
try:
    wbm = load_workbook(MARG, data_only=True, read_only=True)
    wsm = wbm["Neighb_Toronto_ON-Marg2021"]
    # Columns: 0=NH_ID, 1=NH_Name, 2=Pop, then score/quintile pairs:
    # 3/4 households-dwellings, 5/6 material-resources, 7/8 age-labourforce,
    # 9/10 racialized-newcomer. We keep the quintiles (1-5).
    for r in wsm.iter_rows(min_row=4, values_only=True):
        hid = r[0]
        if not isinstance(hid, (int, float)) or not (1 <= int(hid) <= 158):
            continue

        def q(c):
            v = r[c]
            return int(v) if isinstance(v, (int, float)) else None

        marg[int(hid)] = {
            "marg_households": q(4),
            "marg_material": q(6),
            "marg_age_labour": q(8),
            "marg_racialized": q(10),
        }
except Exception as e:  # marginalization is optional; never block the build
    print(f"WARN: marginalization skipped ({e})")


def pct(part, whole, num, nd=1):
    p, w = part.get(num), whole.get(num)
    if isinstance(p, (int, float)) and isinstance(w, (int, float)) and w:
        return round(100.0 * p / w, nd)
    return None


# --- merge into geojson ----------------------------------------------------
with open(GEO, encoding="utf-8") as f:
    gj = json.load(f)

filled = {k: 0 for k in ("pop", "low_income_pct", "transit_commute_pct", "senior_pct", "marg_material")}
for feat in gj["features"]:
    p = feat["properties"]
    num = None
    for k in ("AREA_SHORT_CODE", "AREA_LONG_CODE"):
        if p.get(k) is not None:
            try:
                num = int(p[k])
                break
            except (ValueError, TypeError):
                pass
    name = p.get("AREA_NAME")
    classification = p.get("CLASSIFICATION") or ""
    area = geom_area_km2(feat["geometry"])
    pp = pop.get(num)

    out = {
        "num": num,
        "name": name,
        "is_nia": "Improvement" in classification,
        "area_km2": round(area, 4),
        "pop": pp,
        "density": round(pp / area, 1) if (pp and area) else None,
        "low_income_pct": low_income.get(num),
        "transit_commute_pct": pct(transit, cm_total, num),
        "car_pct": pct(car, cm_total, num),
        "active_pct": (
            round((pct(walked, cm_total, num) or 0) + (pct(bicycle, cm_total, num) or 0), 1)
            if cm_total.get(num) else None
        ),
        "senior_pct": pct(senior, pop, num),
        "renter_pct": pct(renter, ten_total, num),
    }
    for d in range(10):
        out[f"noc{d}_pct"] = pct(noc[d], all_occ, num)
    out.update(marg.get(num, {}))

    p.clear()
    p.update(out)
    round_geom(feat["geometry"], 5)
    for k in filled:
        if out.get(k) is not None:
            filled[k] += 1

with open(OUT, "w", encoding="utf-8") as f:
    json.dump(gj, f, separators=(",", ":"))

n = len(gj["features"])
print(f"wrote {OUT}: {n} features")
print("non-null coverage:", {k: f"{v}/{n}" for k, v in filled.items()})
s = next(x["properties"] for x in gj["features"] if x["properties"]["num"] == 1)
print("sample hood #1:", json.dumps(s, indent=None)[:500])
